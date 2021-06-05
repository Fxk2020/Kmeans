import time

import openpyxl
import pylab
from sklearn import decomposition
from sklearn.cluster import KMeans

import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np

import xlrd
import xlwt
from pylab import mpl

import sys


def writeExcel(path, sheet_name, variablesNames, labels, centers):
    index = len(variablesNames)  # 获取需要写入数据的行数
    index_centers = len(centers)  # 聚类中心数

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    # 向表格中写入数据（对应的行和列）
    outws.cell(1, 1).value = '实例名称'
    for i in range(1, index+1):
        outws.cell(i+1, 1).value = variablesNames[i-1]
    outws.cell(1, 2).value = "对应类别"
    for i in range(1, index+1):
        outws.cell(i+1, 2).value = str(labels[i-1])
    outws.cell(1, 3).value = "聚类中心分别是："
    for i in range(1, index_centers+1):
        outws.cell(i+1, 3).value = str(centers[i-1])
    # for row in range(1,700):
    #     for col in range(1,4):
    #         outws.cell(row, col).value = row*2  # 写文件
    #     print(row)
    saveExcel = path+"/result.xlsx"
    outwb.save(saveExcel)  # 一定要记得保存


def write_excel_xls(path, sheet_name, variablesNames, labels, centers):
    """
    用于将分析结果写入文件
    :param path: 输出路径
    :param sheet_name: 创建的表名
    :param variablesNames: 变量名
    :param labels: 标签
    :param centers: 聚类中心
    :return: 成功创建表格
    """
    index = len(variablesNames)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格

    index_centers = len(centers)

    # 向表格中写入数据（对应的行和列）
    sheet.write(0, 0, "实例名称")
    for i in range(0, index):
        sheet.write(i + 1, 0, variablesNames[i])
    sheet.write(0, 1, "对应类别")
    for i in range(0, index):
        sheet.write(i + 1, 1, str(labels[i]))
    sheet.write(0, 2, "聚类中心分别是：")
    for i in range(0, index_centers):
        sheet.write(i + 1, 2, str(centers[i]))
    # 保存工作簿
    workbook.save(path + "/result.xls")
    print("xls格式表格写入数据成功！")


def loadData(url):
    """
    :param url:待分析文件的路径
    :return: Variable变量名的数组 values变量值的数组
    """

    # 指定默认字体
    # 在mac系统下中文显示
    mpl.rcParams['font.sans-serif'] = ['Arial Unicode MS']
    mpl.rcParams['axes.unicode_minus'] = False  # 解决保存图像是负号'-'显示为方块的问题

    data = xlrd.open_workbook(url)
    table = data.sheet_by_name('data')

    variables = []
    values = np.zeros((table.nrows - 1, table.ncols - 1))
    for row in range(1, table.nrows):
        row_value = []
        variables.append(table.cell_value(row, 0))
        for col in range(1, table.ncols):
            values[row - 1, col - 1] = table.cell_value(row, col)

    variables = np.array(variables)
    # print(values)
    return values, variables


def kmeansMain(values, variables, k, outputDir):
    """
    对输入数据进行聚类分析
    :param values: 聚类的数据
    :param variables: 聚类数据的名称
    :param k: 聚类的个数
    :param outputDir: 图像的输出路径
    :return 聚类数据的标签和聚类中心点
    """
    # 创建模型，进行聚类
    kmeans = KMeans(n_clusters=k)
    kmeans.fit(values)

    # 获取中心点和标签
    centers = kmeans.cluster_centers_
    labels = kmeans.labels_
    PCA(X=values, label=labels, variablesName=variables, outputDir=outputDir)
    return labels, centers


def PCA(X, label, variablesName, outputDir):
    """
    根据两个最大的主成分进行绘图 降维为2，方便画图,输出图像并保存
    :param X: 聚类的源数据
    :param label: 聚完类之后的标签
    :param variablesName:变量名
    :return:直观的聚类结果图像
    """
    pca = decomposition.PCA(n_components=2)
    pca.fit(X)  # 主城分析时每一行是一个输入数据
    result = pca.transform(X)  # 计算结果
    plt.figure(figsize=[10, 6])  # 新建一张图进行绘制
    n_clusters = len(set(label.tolist()))
    for i in range(result[:, 0].size):
        color = cm.nipy_spectral(float(label[i]) / n_clusters)
        plt.plot(result[i, 0], result[i, 1],
                 c=color, marker='o', markersize=10)
        plt.text(result[i, 0], result[i, 1], variablesName[i])
    x_label = 'PC1(%s%%)' % round((pca.explained_variance_ratio_[0] * 100.0), 2)  # x轴标签字符串
    y_label = 'PC1(%s%%)' % round((pca.explained_variance_ratio_[1] * 100.0), 2)  # y轴标签字符串
    plt.xlabel(x_label)  # 绘制x轴标签
    plt.ylabel(y_label)  # 绘制y轴标签
    plt.title('使用主成分分析法对高维数据进行降维，产生直观图像')
    # 显示并保存散点图
    tick = time.time()
    print("当前的时间戳为：", tick)
    pylab.savefig(outputDir + '/result.png')


def main(k, inputUrl, outputUrl):
    """
    脚本的主入口
    :param k: 聚类的个数
    :param inputUrl: 分析文件的url
    :param outputUrl: 结果输出url
    :return:
    """
    values, variables = loadData(inputUrl)
    labels, centers = kmeansMain(values, variables, k=k, outputDir=outputUrl)
    # write_excel_xls(outputUrl, "result", variables, labels, centers)  能写入的数据太少
    writeExcel(outputUrl, "result", variables, labels, centers)


if __name__ == '__main__':
    start = time.perf_counter()
    a = []
    # 其中sys.argv用于获取参数url1，url2等。而sys.argv[0]代表python程序名，所以列表从1开始读取参数。
    for i in range(1, len(sys.argv)):  # 一定要引入sys包！！！！！！
        a.append((sys.argv[i]))
    print(main(int(a[0]), a[1], a[2]))
    end = time.perf_counter()
    print("程序的运行时间为：" + str(end - start))

# main(6,"/Users/yuanbao/Desktop/kmeans算法/data/data.xlsx","/Users/yuanbao/Desktop")
# # writeExcel()
