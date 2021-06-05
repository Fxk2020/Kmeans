import time

import openpyxl
import pylab
from sklearn.mixture import GaussianMixture
from sklearn import decomposition

import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np

import xlrd
import xlwt
from pylab import mpl

import sys


def loadData(url):
    """
    :param url:待分析文件的路径
    :return: value_name实例变量名 x,y数据点的横纵坐标
    """
    # 指定默认字体
    # 在mac系统下中文显示
    mpl.rcParams['font.sans-serif'] = ['Arial Unicode MS']
    mpl.rcParams['axes.unicode_minus'] = False  # 解决保存图像是负号'-'显示为方块的问题

    data = xlrd.open_workbook(url)
    table = data.sheet_by_name('data')

    variables = []
    values = np.zeros((table.nrows - 1, table.ncols - 1))
    for row in range(1, table.nrows):
        row_value = []
        variables.append(table.cell_value(row, 0))
        for col in range(1, table.ncols):
            values[row - 1, col - 1] = table.cell_value(row, col)

    variables = np.array(variables)
    # print(values)
    return values, variables


def write_excel_xls(path,sheet_name,variablesNames,labels):
    """
    将聚类结果写入文件
    :param path: 输出路径
    :param sheet_name: 表名
    :param variablesNames:变量名
    :param labels: 预测的类别
    :return: 成功写入文件
    """
    index = len(variablesNames)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格

    # 向表格中写入数据（对应的行和列）
    sheet.write(0, 0, "实例名称")
    for i in range(0, index):
        sheet.write(i + 1, 0, variablesNames[i])
    sheet.write(0, 1, "对应类别")
    for i in range(0, index):
        sheet.write(i + 1, 1, str(labels[i]))
    # 保存工作簿
    workbook.save(path + "/result.xls")
    print("xls格式表格写入数据成功！")


def writeExcel(path,sheet_name,variablesNames,labels):
    """
        将聚类结果写入文件
        :param path: 输出路径
        :param sheet_name: 表名
        :param variablesNames:变量名
        :param labels: 预测的类别
        :return: 成功写入文件
    """
    index = len(variablesNames)  # 获取需要写入数据的行数
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    # 向表格中写入数据（对应的行和列）
    outws.cell(1, 1).value = '实例名称'
    for i in range(1, index+1):
        outws.cell(i+1, 1).value = variablesNames[i-1]
    outws.cell(1, 2).value = "对应类别"
    for i in range(1, index+1):
        outws.cell(i+1, 2).value = str(labels[i-1])
    saveExcel = path+"/result.xlsx"
    outwb.save(saveExcel)  # 一定要记得保存


def gmm(values, variables,k_number, covariance_type, outputDir):
    """
    gmm主程序
    :param values: 变量
    :param variables: 变量名
    :param k_number: 聚类中心数
    :param covariance_type: 使用的协方差方法
    :param outputDir: 输出路径
    :return: 每个数据的预测类别
    """
    # GaussianMixture 方法中自带了不同的选项来约束不同估类的协方差：spherical，diagonal，tied 或 full 协方差。
    # 'spherical', 'diag', 'tied', 'full'

    # 聚类数据和聚类个数
    X_train = values
    n_classes = k_number

    estimators = GaussianMixture(n_components=n_classes,
                                 covariance_type=covariance_type, max_iter=20, random_state=0)

    estimators.fit(X_train)
    y_predict = estimators.predict(X_train)
    # print(y_predict)
    if len(variables) < 2000:
        PCA(X=values, label=y_predict, variablesName=variables, outputDir=outputDir)
    return y_predict


def PCA(X, label, variablesName, outputDir):
    """
    根据两个最大的主成分进行绘图 降维为2，方便画图,输出图像并保存
    :param X: 聚类的源数据
    :param label: 聚完类之后的标签
    :param variablesName:变量名
    :return:直观的聚类结果图像
    """
    pca = decomposition.PCA(n_components=2)
    pca.fit(X)  # 主城分析时每一行是一个输入数据
    result = pca.transform(X)  # 计算结果
    plt.figure(figsize=[10, 6])  # 新建一张图进行绘制
    n_clusters = len(set(label.tolist()))
    for i in range(result[:, 0].size):
        color = cm.nipy_spectral(float(label[i]) / n_clusters)
        plt.plot(result[i, 0], result[i, 1],
                 c=color, marker='o', markersize=10)
        plt.text(result[i, 0], result[i, 1], variablesName[i])
    x_label = 'PC1(%s%%)' % round((pca.explained_variance_ratio_[0] * 100.0), 2)  # x轴标签字符串
    y_label = 'PC1(%s%%)' % round((pca.explained_variance_ratio_[1] * 100.0), 2)  # y轴标签字符串
    plt.xlabel(x_label)  # 绘制x轴标签
    plt.ylabel(y_label)  # 绘制y轴标签
    plt.title('使用主成分分析法对高维数据进行降维，产生直观图像')
    # 显示并保存散点图
    tick = time.time()
    print("当前的时间戳为：", tick)
    pylab.savefig(outputDir + '/result.png')


def main(covariance_type, k, inputUrl, outputUrl):
    """
    脚本入口
    :param k: 聚类个数
    :param inputUrl: 输入路径
    :param outputUrl: 输出路径
    :return:
    """
    values, variables = loadData(inputUrl)
    labels = gmm(values, variables, k_number=k, covariance_type=covariance_type, outputDir=outputUrl)
    # write_excel_xls(outputUrl,"result",variables,labels)
    writeExcel(outputUrl,'result',variables,labels)


# main("full", 4,"/Users/yuanbao/Desktop/kmeans算法/data/data.xlsx","/Users/yuanbao/Desktop")

if __name__ == '__main__':
    start = time.perf_counter()
    a = []
    # 其中sys.argv用于获取参数url1，url2等。而sys.argv[0]代表python程序名，所以列表从1开始读取参数。
    for i in range(1, len(sys.argv)):  # 一定要引入sys包！！！！！！
        a.append((sys.argv[i]))
    main(a[0], int(a[1]), a[2], a[3])
    end = time.perf_counter()
    print("程序的运行时间为：" + str(end - start))
